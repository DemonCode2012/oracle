from openpyxl import Workbook
import datetime

workDesc = 'ORACLE驻场运维'
workType = '现场技术处理'
beginList = []
endList = []
# today = datetime.date.today()
# print(today)
dinput = input("请输入日期 格式yyyymmdd：\n")
date = datetime.datetime.strptime(dinput, '%Y%m%d')
print(date.date())
#
#
for i in range(5):
    # day = today + datetime.timedelta(i - today.weekday())
    day = date.date() + datetime.timedelta(i - date.weekday())
    # print(day)
    # beginList.append(str(day.strftime("%Y/%m/%d")) + ' 09:00:00')
    # endList.append(str(day.strftime("%Y/%m/%d")) + ' 17:00:00')
    beginList.append(str(day) + ' 09:30:00')
    endList.append(str(day) + ' 18:00:00')
print(beginList)
print(endList)

wb = Workbook()
ws = wb.active

ws['A1'] = '开始时间'
ws['B1'] = '结束时间'
ws['C1'] = '工作描述'
ws['D1'] = '工时类型'

for i in range(5):
    ws['A' + str(i + 2)] = datetime.datetime.strptime(beginList[i], '%Y-%m-%d %H:%M:%S')
    # ws['A' + str(i + 2)].number_format = AFormat
    ws['B' + str(i + 2)] = datetime.datetime.strptime(endList[i], '%Y-%m-%d %H:%M:%S')
    # ws['B' + str(i + 2)].number_format = BFormat
    # ws['C' + str(i + 2)].number_format = CFormat
    # ws['D' + str(i + 2)].number_format = DFormat
    # ws['A' + str(i + 2)] = beginList[i]
    # ws['B' + str(i + 2)] = endList[i]
    ws['C' + str(i + 2)] = workDesc
    ws['D' + str(i + 2)] = workType
    print(beginList[i], endList[i], workDesc, workType)
monday = (date.date() + datetime.timedelta(0 - date.weekday())).strftime('%Y%m%d')
sunday = (date.date() + datetime.timedelta(6 - date.weekday())).strftime('%m%d')
print(r'D:\1.Project\CES\2.售前工时\CES工时导入-'+monday+'-'+sunday+'.xlsx')
wb.save(r'D:\1.Project\CES\2.售前工时\CES工时导入-'+monday+'-'+sunday+'.xlsx')
